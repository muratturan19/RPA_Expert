# -*- coding: utf-8 -*-
"""POS entry automation script with UTF-8/Turkish support."""

import argparse
import logging
import sys
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Any

from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service as ChromeService
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import WebDriverException, TimeoutException
from webdriver_manager.chrome import ChromeDriverManager


if sys.stdout.encoding != "utf-8":
    sys.stdout.reconfigure(encoding="utf-8")
if sys.stderr.encoding != "utf-8":
    sys.stderr.reconfigure(encoding="utf-8")

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.FileHandler("rpa.log", encoding="utf-8"),
        logging.StreamHandler(sys.stdout)
    ]
)
logger = logging.getLogger(__name__)


def read_excel(path: Path) -> List[Dict[str, Any]]:
    """Read POS data from an Excel file."""
    logger.info("Reading Excel data from %s", path)
    workbook = load_workbook(path)
    sheet = workbook.active
    rows: List[Dict[str, Any]] = []

    column_map = {
        "tarih": "tarih",
        "firma": "firma",
        "tutar": "tutar",
        "açıklama": "aciklama",
        "aciklama": "aciklama",
        "döviz": "doviz",
        "doviz": "doviz",
        "vade tarihi": "vade_tarihi",
    }

    raw_header = [str(cell).strip() if cell else "" for cell in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
    header = [column_map.get(h.casefold(), h) for h in raw_header]

    for row in sheet.iter_rows(min_row=2, values_only=True):
        item: Dict[str, Any] = {}
        for key, value in zip(header, row):
            if isinstance(value, datetime):
                item[key] = value.date().isoformat()
            else:
                item[key] = value
        rows.append(item)
    logger.info("Loaded %d rows from Excel", len(rows))
    return rows


def setup_driver() -> webdriver.Chrome:
    """Initialise Chrome WebDriver."""
    options = Options()
    options.add_argument("--start-maximized")
    driver = webdriver.Chrome(service=ChromeService(ChromeDriverManager().install()), options=options)
    return driver


def open_application(driver: webdriver.Chrome, html_path: Path) -> None:
    """Open local Preston simulator."""
    driver.get(html_path.resolve().as_uri())
    logger.info("Opened Preston simulator at %s", html_path)


def _ensure_overlay_closed(driver: webdriver.Chrome, timeout: int = 10) -> None:
    """Ensure that loading or modal overlays are not blocking interactions."""
    try:
        WebDriverWait(driver, timeout).until(
            lambda d: not d.execute_script("return document.body.classList.contains('loading')")
        )
        WebDriverWait(driver, timeout).until(
            EC.invisibility_of_element_located((By.ID, "modalOverlay"))
        )
    except TimeoutException:
        driver.execute_script(
            "document.body.classList.remove('loading');"
            "var o=document.getElementById('modalOverlay');"
            "if(o){o.style.display='none';}"
        )


def navigate_to_pos(driver: webdriver.Chrome) -> None:
    """Navigate to Finance > Tahsilat and open POS entry modal."""
    menu_selector = "div.menu-item[data-menu='finans-tahsilat']"
    pos_icon_selector = (
        "div.ribbon-icon[data-tooltip='Pos Girişi'],"
        "div.ribbon-icon[data-tooltip='POS Girişi']"
    )
    _ensure_overlay_closed(driver)
    WebDriverWait(driver, 10).until(
        EC.element_to_be_clickable((By.CSS_SELECTOR, menu_selector))
    ).click()
    _ensure_overlay_closed(driver)
    WebDriverWait(driver, 10).until(
        EC.element_to_be_clickable((By.CSS_SELECTOR, ", ".join(pos_icon_selector)))
    ).click()
    WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.ID, "posModal")))
    logger.debug("POS modal opened")


def fill_pos_form(driver: webdriver.Chrome, data: Dict[str, Any]) -> None:
    """Fill POS entry form and save."""
    driver.find_element(By.ID, "posTarih").clear()
    driver.find_element(By.ID, "posTarih").send_keys(str(data.get("tarih", "")))

    driver.find_element(By.ID, "posKartHesap").send_keys(str(data.get("firma", "")))
    driver.find_element(By.ID, "posAciklama").send_keys(str(data.get("aciklama", "")))
    driver.find_element(By.ID, "posTutar").send_keys(str(data.get("tutar", "")))

    doviz = data.get("doviz")
    if doviz:
        driver.find_element(By.ID, "posDoviz").send_keys(str(doviz))

    vade = data.get("vade_tarihi")
    if vade:
        driver.find_element(By.ID, "posVadeTarihi").send_keys(str(vade))

    driver.find_element(By.CSS_SELECTOR, ".modal-buttons .primary").click()
    _ensure_overlay_closed(driver)
    logger.info("POS entry saved for %s", data)


def process_entries(driver: webdriver.Chrome, entries: List[Dict[str, Any]]) -> None:
    for entry in entries:
        for attempt in range(3):
            try:
                navigate_to_pos(driver)
                fill_pos_form(driver, entry)
                _ensure_overlay_closed(driver)
                break
            except Exception as exc:  # pylint: disable=broad-except
                logger.exception(
                    "Error processing row %s on attempt %d: %s", entry, attempt + 1, exc
                )
                _ensure_overlay_closed(driver)
        else:
            logger.error("Failed to process row %s after retries", entry)


def main(excel_path: Path) -> None:
    data_rows = read_excel(excel_path)
    driver = setup_driver()
    try:
        html_file = Path(__file__).parent / "RPA_Expert.html"
        open_application(driver, html_file)
        process_entries(driver, data_rows)
    except WebDriverException as exc:
        logger.exception("WebDriver error: %s", exc)
    finally:
        driver.quit()
        logger.info("Driver closed")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Automate POS entries on Preston simulator")
    parser.add_argument("--excel", default="pos_data.xlsx", help="Path to Excel file with POS data")
    args = parser.parse_args()
    main(Path(args.excel))
